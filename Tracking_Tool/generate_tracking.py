import os
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook

# ==========================================
# 1. 文件夹配置
# ==========================================
INPUT_DIR = "input_pi"          # 读取生成好的PI文件夹（包含01, 02等子文件夹）
OUTPUT_DIR = "output_tracking"    # 明确输出到你指定的 output_tracking 文件夹

def generate_reports():
    # 确保目标输出文件夹存在，如果没有会自动创建
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    
    print(">>> 正在扫描所有子文件夹中的 PI 文件...")
    
    # 自动穿透子文件夹，寻找所有 Excel 文件
    excel_filepaths = []
    for root, dirs, files in os.walk(INPUT_DIR):
        for file in files:
            # 排除临时文件
            if file.endswith('.xlsx') and not file.startswith('~') and not file.startswith('Batch_'):
                excel_filepaths.append(os.path.join(root, file))
                
    excel_filepaths.sort() 
    
    if not excel_filepaths:
        print(f"【提示】没有在 [{INPUT_DIR}] 文件夹中找到可以提取的 PI Excel 文件，请确认路径。")
        return

    tracking_records = []

    for filepath in excel_filepaths:
        print(f"  正在提取 -> {os.path.basename(filepath)}")
        
        # data_only=True 确保读取的是人工核对修改后的最终数值，而不是公式本身
        wb = load_workbook(filepath, data_only=True)
        ws = wb.active
        
        # --- 提取表头固定信息 ---
        pi_number = ws["F8"].value or ""
        po_number = ws["F9"].value or ""
        
        # 核心修正：原封不动保留多行完整地址（含换行符，绝不漏掉电话、邮箱和原本的排版）
        bill_to_raw = str(ws["B8"].value or "").strip()
        ship_to_raw = str(ws["B10"].value or "").strip()

        # --- 提取动态产品行 ---
        current_row = 15
        while True:
            sku = ws[f"B{current_row}"].value
            if not sku:  # B列（SKU）为空代表产品读取完毕
                break
                
            # 完整抓取每一列的数据
            item_no = ws[f"A{current_row}"].value or ""     # Item 序号
            sku_name = ws[f"C{current_row}"].value or ""    # 产品的英文名
            
            # 强制转换为数字进行安全计算
            try:
                qty = float(ws[f"D{current_row}"].value or 0)
                price = float(ws[f"E{current_row}"].value or 0)
            except ValueError:
                qty = 0.0
                price = 0.0
                
            # 该带计算的自动带计算：用 Python 重新核算总价，防止 Excel 公式意外丢失
            calculated_amount = qty * price

            # 严格按照你指定的字段和顺序组装 Tracking 数据
            tracking_records.append({
                "Item": item_no,
                "PO number": po_number,
                "PI number": pi_number,
                "Billing To": bill_to_raw,
                "Delivery Address": ship_to_raw,
                "SKU Number": sku,
                "SKU Name": sku_name,
                "Order Qty": qty,
                "Price": price,
                "Total Amount": calculated_amount
            })

            current_row += 1 

    # ==========================================
    # 2. 导出到目标文件夹
    # ==========================================
    if tracking_records:
        df_tracking = pd.DataFrame(tracking_records)
        run_time = datetime.now().strftime("%y%m%d_%H%M%S")
        export_filename = os.path.join(OUTPUT_DIR, f"Batch_Tracking_{run_time}.xlsx")
        
        # 导出为干净的目标大表
        df_tracking.to_excel(export_filename, index=False)
        print(f"\n🎉 成功！Tracking大表已完美生成在目标文件夹: {export_filename}")

if __name__ == "__main__":
    generate_reports()