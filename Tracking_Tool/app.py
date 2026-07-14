import streamlit as st
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from io import BytesIO

# ==========================================
# 1. 网页 UI 设置
# ==========================================
st.set_page_config(page_title="PI Tracking 汇总工具", page_icon="📊", layout="centered")
st.title("📊 PI Tracking 自动汇总工具")
st.markdown("请将检查无误的 **PI Excel 文件**拖拽到下方区域（支持一次性拖入多个文件）。")

# 拖拽上传组件
uploaded_files = st.file_uploader("拖拽 Excel 文件到这里", type=['xlsx'], accept_multiple_files=True)

# ==========================================
# 2. 核心提取逻辑 (100% 对齐你的 Python 版本)
# ==========================================
if st.button("🚀 一键生成 Tracking 大表") and uploaded_files:
    with st.spinner('正在光速处理中...'):
        tracking_records = []

        for uploaded_file in uploaded_files:
            # data_only=True 确保读取的是最终数值
            wb = load_workbook(uploaded_file, data_only=True)
            ws = wb.active
            
            # --- 提取表头 ---
            pi_number = ws["F8"].value or ""
            po_number = ws["F9"].value or ""
            
            # 核心保留：原封不动保留多行完整地址
            bill_to_raw = str(ws["B8"].value or "").strip()
            ship_to_raw = str(ws["B10"].value or "").strip()

            # --- 提取动态产品行 ---
            current_row = 15
            while True:
                sku = ws[f"B{current_row}"].value
                if not sku:
                    break
                    
                item_no = ws[f"A{current_row}"].value or ""     
                sku_name = ws[f"C{current_row}"].value or ""    
                
                try:
                    qty = float(ws[f"D{current_row}"].value or 0)
                    price = float(ws[f"E{current_row}"].value or 0)
                except ValueError:
                    qty = 0.0
                    price = 0.0
                    
                # Python 重新核算总价
                calculated_amount = qty * price

                # 严格按照你指定的字段组装
                tracking_records.append({
                    "Item": item_no,
                    "PO number": po_number,
                    "PI number": pi_number,
                    "Billing To": bill_to_raw,
                    "Delivery Address": ship_to_raw,
                    "SKU Number": sku,
                    "SKU Name": sku_name,
                    "Order Qty": qty,
                    "Price": price,
                    "Total Amount": calculated_amount
                })

                current_row += 1 

        # ==========================================
        # 3. 内存导出与下载 (不产生本地垃圾文件)
        # ==========================================
        if tracking_records:
            # 使用 BytesIO 在内存中直接生成 Excel，无需写在电脑硬盘上
            output = BytesIO()
            df_tracking = pd.DataFrame(tracking_records)
            df_tracking.to_excel(output, index=False)
            excel_data = output.getvalue()
            
            st.success("✅ 处理完成！请点击下方按钮下载。")
            
            # 下载按钮
            run_time = datetime.now().strftime("%y%m%d_%H%M%S")
            st.download_button(
                label="⬇️ 下载 Batch_Tracking 汇总表",
                data=excel_data,
                file_name=f"Batch_Tracking_{run_time}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.warning("⚠️ 没有在文件中提取到有效的产品数据。")