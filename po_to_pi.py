import os
import re
import shutil
from copy import copy
from datetime import datetime
import pdfplumber
import pandas as pd
from openpyxl import load_workbook

# ==========================================
# 1. 配置文件路径与表格坐标
# ==========================================
TEMPLATE_PATH = "templates/PI_template.xlsx"
SKU_TABLE_PATH = "data/internal_sku.xlsx"
INPUT_DIR = "input_po"
OUTPUT_DIR = "output_pi"
COUNTER_FILE = "data/pi_counter.txt"

CELL_PI_NO = "F8"      
CELL_PO_NO = "F9"      
CELL_INVOICE_DATE = "F10"  
CELL_BILL_TO = "B8"    
CELL_SHIP_TO = "B10"   

START_ROW_PRODUCTS = 15  
COL_ITEM = "A"         
COL_SKU = "B"          
COL_PROD = "C"         
COL_QTY = "D"          
COL_PRICE = "E"        
COL_AMOUNT = "F"       

# ==========================================
# 2. 核心提取雷达 (加入强力去重机制)
# ==========================================
def parse_products_from_text(text_normal):
    """Parse product blocks from the rows that contain quantity and pricing."""
    products, seen_features = [], set()
    lines = [line.strip() for line in text_normal.splitlines() if line.strip()]
    money_row = re.compile(r"\b(\d+)\s+\$[\d,.]+\s+(?:\$[\d,.]+\s+)?\$?[\d,.]+\s*$")
    item_words = ("omni", "footrest", "wheel", "stepsync", "battery", " mat", "seat pad")

    for i, line in enumerate(lines):
        qty_match = money_row.search(line)
        if not qty_match or "total" in line.lower():
            continue
        block = [line]
        for next_line in lines[i + 1:i + 6]:
            if money_row.search(next_line) or "purchase order total" in next_line.lower() or next_line.lower().startswith("total:"):
                break
            block.append(next_line)
        raw_name = " ".join(block)
        normalized = clean_text_for_match(raw_name)
        if not any(word in normalized for word in item_words):
            continue
        # Pricing columns and source SKU must not make the same product unique.
        feature = re.sub(r"\b(?:cp|cb)\s+\w+(?:\s+\w+){0,5}\b", "", normalized)
        feature = re.sub(r"\b\d+\s+\d+(?:\s+\d+){1,3}\b", "", feature)
        feature = re.sub(r"\s+", " ", feature).strip()
        if feature in seen_features:
            continue
        seen_features.add(feature)

        products.append({"type": product_type_from_text(normalized), "raw_name": raw_name, "qty": int(qty_match.group(1))})

    # Footrests frequently omit the chair size. Infer it only from a unique
    # chair with the same family and colour in this PO.
    return enrich_product_features(products)

def product_type_from_text(normalized):
    # This is a chair configuration, not a separately billable footrest line.
    if "chair with footrest" in normalized:
        return "main_chair"
    if "footrest" in normalized:
        return "footrest"
    if "wheel" in normalized:
        return "accessory"
    if "battery" in normalized:
        return "battery"
    if "stepsync mat" in normalized or re.search(r"\bmat\b", normalized):
        return "mat"
    if "seat pad" in normalized:
        return "seat_pad"
    return "main_chair"

def enrich_product_features(products):
    """Add an unambiguous chair size to footrests that omit it in their row."""
    expanded_products = []
    for product in products:
        expanded_products.append(product)
        normalized = clean_text_for_match(product["raw_name"])
        if product["type"] == "main_chair" and "chair with footrest" in normalized:
            # This legacy PO sells the chair as one line, but the PI requires
            # the included footrest on its own zero-price line.
            expanded_products.append({
                "type": "footrest",
                "raw_name": product["raw_name"],
                "qty": product["qty"],
                "included_footrest": True,
            })
    products = expanded_products

    chairs = {}
    for product in products:
        if product["type"] == "main_chair":
            family, color, length = extract_family_color_length(product["raw_name"])
            if family and color and length:
                chairs.setdefault((family, color), set()).add(length)
    for product in products:
        if product["type"] == "footrest":
            family, color, length = extract_family_color_length(product["raw_name"])
            candidates = chairs.get((family, color), set())
            if family and color and not length and len(candidates) == 1:
                product["inferred_length"] = next(iter(candidates))
    return products

def parse_products_from_pages(pages, text_normal):
    """Read columnar PO tables first; fall back to text for non-tabular formats."""
    products = []
    code_start = re.compile(r"(?=(?:CP-[A-Z0-9]+-[A-Z0-9-]+|CBO\d-[A-Z0-9-]+))", re.IGNORECASE)

    for page in pages:
        for table in page.extract_tables():
            header_row = next((row for row in table if any(str(cell or '').strip().lower() == 'description' for cell in row)), None)
            if not header_row:
                continue
            description_col = next(i for i, cell in enumerate(header_row) if str(cell or '').strip().lower() == 'description')
            quantity_col = next((i for i, cell in enumerate(header_row) if str(cell or '').strip().lower() == 'quantity'), None)
            if quantity_col is None:
                continue
            header_index = table.index(header_row)
            descriptions = "\n".join(str(row[description_col] or '') for row in table[header_index + 1:] if len(row) > description_col)
            quantities = []
            for row in table[header_index + 1:]:
                if len(row) > quantity_col:
                    quantities.extend(int(value) for value in re.findall(r"\b\d+\b", str(row[quantity_col] or '')))
            blocks = [part.strip() for part in code_start.split(descriptions) if part.strip()]
            if not blocks or len(blocks) != len(quantities):
                continue
            for raw_name, qty in zip(blocks, quantities):
                normalized = clean_text_for_match(raw_name)
                # A source SKU is an item boundary. Keep unknown future
                # categories instead of silently dropping them.
                products.append({"type": product_type_from_text(normalized), "raw_name": raw_name, "qty": qty})

    if not products:
        return parse_products_from_text(text_normal)
    return enrich_product_features(products)

def extract_po_number(text_normal, patterns):
    for pattern in patterns:
        match = re.search(pattern, text_normal, re.IGNORECASE | re.MULTILINE)
        if match:
            po = match.group(1).strip()
            po = re.sub(r'(?i)^.*?\b(\d+[A-Za-z0-9\-]*)\b.*$', r'\1', po)
            po = re.sub(r'\s+', ' ', po)
            po = re.sub(r'(?i)\b(date raised|reference|supplier|invoice to|deliver to|invoice date|ship to)\b.*$', '', po).strip()
            po = po.rstrip(':-，,;')
            if po:
                return po
    return "UNKNOWN_PO"

# ==========================================
# 3. 核心数据提取路由 
# ==========================================
def extract_po_format_A(page, text_normal):
    """【老格式】"""
    info = {"po_number": "", "bill_to": "⚠️ 地址定位失败", "deliver_to": "⚠️ 地址定位失败", "products": []}
    
    info["po_number"] = extract_po_number(
        text_normal,
        [
            r"(?:PURCHASE\s*ORDER(?:\s*REPRINT)?)\s+([0-9A-Za-z\-]+)\s*-",
            r"Purchase Order Number:\s*([^\n\r]+)",
            r"PO\s*#\s*[:：]?\s*([^\n\r]+)",
            r"PO#\s*[:：]?\s*([^\n\r]+)",
            r"Order No\.\s*[:：]?\s*([^\n\r]+)",
            r"PURCHASE\s*ORDER\s*([^\n\r]+)",
        ],
    )
    
    words = page.extract_words()
    bill_x0, deliver_x0 = None, None
    address_top, address_bottom = None, page.height
    
    for i, w in enumerate(words):
        if w['text'] in ['Bill', 'Invoice'] and i+1 < len(words) and 'to:' in words[i+1]['text']:
            bill_x0 = w['x0']
            address_top = w['bottom'] 
        if w['text'] in ['Deliver', 'Ship'] and i+1 < len(words) and 'to:' in words[i+1]['text']:
            deliver_x0 = w['x0']
            
    if address_top:
        for w in words:
            if w['text'] in ['EORI', 'SKU', 'Quantity']:
                if w['top'] > address_top and w['top'] < address_bottom:
                    address_bottom = w['top'] - 2
                    
    if bill_x0 and deliver_x0 and address_top:
        bill_bbox = (max(0, bill_x0 - 2), address_top, deliver_x0 - 5, address_bottom)
        deliver_bbox = (max(0, deliver_x0 - 2), address_top, page.width, address_bottom)
        
        raw_bill = page.crop(bill_bbox).extract_text() or ""
        raw_deliver = page.crop(deliver_bbox).extract_text() or ""
        
        info["bill_to"] = re.sub(r"(?i)(Bill to:|Invoice to:|Deliver to:|Ship to:)", "", raw_bill).strip()
        info["deliver_to"] = re.sub(r"(?i)(Bill to:|Invoice to:|Deliver to:|Ship to:)", "", raw_deliver).strip()

    info["products"] = parse_products_from_text(text_normal)
    return info

def extract_po_format_B(page, text_normal):
    """【新格式】"""
    info = {"po_number": "", "bill_to": "⚠️ 地址定位失败", "deliver_to": "⚠️ 地址定位失败", "products": []}
    info["po_number"] = extract_po_number(
        text_normal,
        [
            r"PO\s*#\s*[:：]?\s*([^\n\r]+)",
            r"PO#\s*[:：]?\s*([^\n\r]+)",
            r"Purchase Order Number:\s*([^\n\r]+)",
            r"Order No\.\s*[:：]?\s*([^\n\r]+)",
            r"(?:PURCHASE\s*ORDER(?:\s*REPRINT)?)\s+([0-9A-Za-z\-]+)\s*-",
        ],
    )

    # Format B has reliable, explicit labels in the extracted text. Prefer this
    # bounded form so the address never absorbs the vendor or the item table.
    flat_text = re.sub(r'\s+', ' ', text_normal)
    bill_match = re.search(r'(?i)BILL TO ADDRESS:\s*(.*?)(?:VENDOR:|SHIP TO LOCATION:)', flat_text)
    ship_match = re.search(r'(?i)SHIP TO LOCATION:\s*(.*?)(?:RECEIVING NOTES:|ORDER DATE|PLU\b|$)', flat_text)
    if bill_match:
        info["bill_to"] = bill_match.group(1).strip()
    if ship_match:
        info["deliver_to"] = ship_match.group(1).strip()

    words = page.extract_words()
    bill_x0 = None
    ship_x0 = None
    header_top = None
    stop_top = page.height
    
    for i, w in enumerate(words):
        t = re.sub(r'[^\w]', '', w['text']).lower()
        if t == 'invoice' and i + 1 < len(words) and re.sub(r'[^\w]', '', words[i + 1]['text']).lower() == 'to':
            bill_x0 = w['x0']
            header_top = w['bottom'] if header_top is None else min(header_top, w['bottom'])
        if t == 'bill' and i + 1 < len(words) and re.sub(r'[^\w]', '', words[i + 1]['text']).lower() == 'to':
            bill_x0 = w['x0']
            header_top = w['bottom'] if header_top is None else min(header_top, w['bottom'])
        if t == 'ship' and i + 1 < len(words) and re.sub(r'[^\w]', '', words[i + 1]['text']).lower() == 'to':
            ship_x0 = w['x0']
            header_top = w['bottom'] if header_top is None else min(header_top, w['bottom'])
        if t == 'deliver' and i + 1 < len(words) and re.sub(r'[^\w]', '', words[i + 1]['text']).lower() == 'to':
            ship_x0 = w['x0']
            header_top = w['bottom'] if header_top is None else min(header_top, w['bottom'])
        if t in ['eori', 'sku']:
            stop_top = min(stop_top, w['top'])
                
    try:
        if bill_x0 is not None and ship_x0 is not None and header_top is not None:
            bill_bbox = (max(0, bill_x0 - 2), header_top, ship_x0 - 4, stop_top)
            raw_bill = (page.crop(bill_bbox).extract_text() or "").strip()
            raw_bill = re.sub(r"(?i)(bill to address:|invoice to address:|invoice to:)", "", raw_bill)
            raw_bill = re.sub(r"(?i)vendor:.*$", "", raw_bill).strip()
            info["bill_to"] = re.sub(r'\s+', ' ', raw_bill).strip()

            ship_bbox = (max(0, ship_x0 - 2), header_top, page.width, stop_top)
            raw_deliver = (page.crop(ship_bbox).extract_text() or "").strip()
            raw_deliver = re.sub(r"(?i)(ship to location:|deliver to:)", "", raw_deliver)
            raw_deliver = re.sub(r"(?i)(receiving notes:|order date|receiving order).*$", "", raw_deliver).strip()
            info["deliver_to"] = re.sub(r'\s+', ' ', raw_deliver).strip()
    except Exception:
        pass

    # The right-hand metadata and left-hand vendor are interleaved in text
    # extraction on this layout. Crop the two address columns by coordinates.
    def matching_word(label):
        return next((word for word in words if word['text'].rstrip(':').lower() == label), None)

    bill_word = matching_word('bill')
    vendor_word = matching_word('vendor')
    ship_word = matching_word('ship')
    receiving_word = matching_word('receiving')
    if bill_word and vendor_word and ship_word and receiving_word:
        def keep_address_lines(raw_address, label_pattern):
            cleaned = re.sub(label_pattern, '', raw_address, flags=re.IGNORECASE).strip()
            return "\n".join(line.strip() for line in cleaned.splitlines() if line.strip())

        bill_bbox = (max(0, bill_word['x0'] - 2), bill_word['bottom'], ship_word['x0'] - 8, vendor_word['top'] - 2)
        raw_bill = (page.crop(bill_bbox).extract_text() or '').strip()
        if raw_bill:
            info['bill_to'] = keep_address_lines(raw_bill, r'^bill to address:\s*')

        location_word = next((word for word in words if word['text'].rstrip(':').lower() == 'location' and abs(word['top'] - ship_word['top']) < 2), None)
        ship_start = (location_word or ship_word)['x1'] + 2
        ship_bbox = (ship_start, ship_word['top'], page.width, receiving_word['top'] - 2)
        raw_ship = (page.crop(ship_bbox).extract_text() or '').strip()
        if raw_ship:
            info['deliver_to'] = keep_address_lines(raw_ship, r'^ship to location:\s*')

    if info["bill_to"] == "⚠️ 地址定位失败" or info["deliver_to"] == "⚠️ 地址定位失败":
        bill_match = re.search(r'(?i)BILL TO\s*[:：]?\s*(.*?)(?:PO\s*#|SHIP TO\s*[:：]?)', flat_text)
        if bill_match:
            block = bill_match.group(1).strip()
            block = re.sub(r'(?i)^INVOICE#\s*:\s*\S+\s*', '', block).strip()
            if block:
                info["bill_to"] = block

        ship_match = re.search(r'(?i)SHIP TO\s*[:：]?\s*(.*?)(?:INVOICE DATE|DUE DATE|Payment Term|Trade Term|No\. Part Number|$)', flat_text)
        if ship_match:
            block = ship_match.group(1).strip()
            if not re.search(r'\d', block):
                street_match = re.search(r'(?i)INVOICE DATE:\s*[0-9]{4}-[0-9]{2}-[0-9]{2}\s+(\d{1,5}\b.*?)(?:DUE DATE|Payment Term|Trade Term|No\. Part Number|$)', flat_text)
                if street_match:
                    block = street_match.group(1).strip()
            if block:
                info["deliver_to"] = block

    info["products"] = parse_products_from_text(text_normal)
    return info

def extract_po_info(pdf_path):
    print(f"\n>>> 正在解析 PDF: {os.path.basename(pdf_path)}")
    with pdfplumber.open(pdf_path) as pdf:
        page = pdf.pages[0]
        # Addresses belong to the first-page header, while line items may span
        # any number of pages. Preserve page boundaries to avoid merged words.
        page_texts = [(p.extract_text() or "") for p in pdf.pages]
        text_normal = "\n".join(page_texts)
        
        text_lower = text_normal.lower()
        if (
            "ship to location" in text_lower
            or "bill to address" in text_lower
            or "invoice to address" in text_lower
            or ("bill to" in text_lower and "ship to" in text_lower)
        ):
            info = extract_po_format_B(page, text_normal)
        else:
            info = extract_po_format_A(page, text_normal)
        info["products"] = parse_products_from_pages(pdf.pages, text_normal)
        return info

# ==========================================
# 4. 智能记忆与匹配 
# ==========================================
def generate_pi_no():
    today_str = datetime.now().strftime("%y%m%d")
    prefix = f"LNN19{today_str}"
    count = 1
    if os.path.exists(COUNTER_FILE):
        with open(COUNTER_FILE, "r") as f:
            content = f.read().strip()
            if content and "," in content:
                saved_prefix, saved_count = content.split(",")
                if saved_prefix == prefix:
                    count = int(saved_count) + 1
    with open(COUNTER_FILE, "w") as f:
        f.write(f"{prefix},{count}")
    return f"{prefix}{count:03d}", count 

df_sku = pd.read_excel(SKU_TABLE_PATH)
df_sku.columns = [str(c).strip() for c in df_sku.columns]
df_sku['kw_len'] = df_sku['PO_Keyword'].astype(str).str.len()
df_sku = df_sku.sort_values('kw_len', ascending=False)

def clean_text_for_match(text):
    t = str(text).lower()
    t = re.sub(r'([a-z])(\d)', r'\1 \2', t) 
    t = re.sub(r'[^\w\s]', ' ', t)          
    return re.sub(r'\s+', ' ', t).strip()   

def extract_family_color_length(text):
    normalized = clean_text_for_match(text)
    family = None
    for candidate in ['omni pro', 'omni gen', 'omni se', 'omni dynamic', 'wheels', 'footrest']:
        if candidate in normalized:
            family = candidate
            break

    color = None
    color_aliases = [
        ('space gray', 'gray'),
        ('midnight black', 'black'),
        ('moss green', 'green'),
        ('gray', 'gray'),
        ('black', 'black'),
        ('green', 'green'),
        ('glacier', 'glacier'),
        ('graphite', 'graphite'),
        ('obsidian', 'obsidian'),
    ]
    for token, mapped in color_aliases:
        if token in normalized:
            color = mapped
            break

    length = None
    length_match = re.search(r'(?<!\d)(45|48)(?!\d)', normalized)
    if length_match:
        length = length_match.group(1)

    return family, color, length

def extract_keyword_signature(keyword_text):
    family, color, length = extract_family_color_length(keyword_text)
    if not family or not color:
        return None
    return family, color, length

def product_category(text, product_type=None):
    normalized = clean_text_for_match(text)
    if product_type == "footrest" or "footrest" in normalized:
        return extract_family_color_length(normalized)[0]
    if product_type == "accessory" or "wheel" in normalized:
        return "wheels"
    if product_type == "battery" or "battery" in normalized:
        return "battery"
    if product_type == "mat" or "stepsync mat" in normalized:
        return "stepsync mat"
    if product_type == "seat_pad" or "seat pad" in normalized:
        return "seat pad"
    return extract_family_color_length(normalized)[0]

def map_product_from_single_table(prod_dict):
    sku, eng_name, price = "需核对", "需核对", 0.0
    raw_name_clean = clean_text_for_match(prod_dict['raw_name'])

    raw_family, raw_color, raw_length = extract_family_color_length(raw_name_clean)
    raw_category = product_category(raw_name_clean, prod_dict["type"])
    raw_length = prod_dict.get("inferred_length", raw_length)
    # CBO1 is the legacy Omni Dynamic 48 cm chair code; its PDF description
    # omits the size but it is not an ambiguous fuzzy product match.
    if raw_length is None and raw_category == "omni dynamic" and re.search(r"\bcbo1-", prod_dict["raw_name"], re.IGNORECASE):
        raw_length = "48"

    for _, row in df_sku.iterrows():
        keyword_str = clean_text_for_match(row.get('PO_Keyword', ''))
        if not keyword_str or keyword_str == 'nan':
            continue

        kw_family, kw_color, kw_length = extract_family_color_length(keyword_str)
        kw_category = product_category(keyword_str)
        if not kw_category or not kw_color:
            continue

        category_match = raw_category == kw_category
        color_match = raw_color == kw_color
        length_match = kw_length is None or raw_length == kw_length

        if category_match and color_match and length_match:
            if prod_dict['type'] == "footrest":
                sku = row.get('脚踏SKU', '需核对')
                eng_name = row.get('脚踏English', '需核对')
                price_val = row.get('脚踏价格', 0)
                price = float(price_val) if pd.notna(price_val) else 0.0
            else:
                sku = row.get('SKU', '需核对')
                eng_name = row.get('产品English', '需核对')
                price_val = row.get('椅子价格', 0) 
                price = float(price_val) if pd.notna(price_val) else 0.0
            break 
    return sku, eng_name, price

def copy_row_style(ws, source_row, target_row, max_col=6):
    if source_row == target_row:
        return
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col_idx in range(1, max_col + 1):
        source_cell = ws.cell(source_row, col_idx)
        target_cell = ws.cell(target_row, col_idx)
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
        target_cell.number_format = source_cell.number_format
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.protection = copy(source_cell.protection)

def refresh_total_row(ws, start_row, end_row, total_row):
    ws[f"A{total_row}"] = "Total amount"
    ws[f"D{total_row}"] = f"=SUM(D{start_row}:D{end_row})"
    ws[f"F{total_row}"] = f"=SUM(F{start_row}:F{end_row})"

def insert_product_rows(ws, first_insert_row, row_count):
    """Insert detail rows without leaving payment-area merged cells behind."""
    if row_count <= 0:
        return
    moved_ranges = [
        (cell_range.min_col, cell_range.min_row, cell_range.max_col, cell_range.max_row)
        for cell_range in ws.merged_cells.ranges
        if cell_range.min_row >= first_insert_row
    ]
    for min_col, min_row, max_col, max_row in moved_ranges:
        ws.unmerge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)

    ws.insert_rows(first_insert_row, row_count)
    for min_col, min_row, max_col, max_row in moved_ranges:
        ws.merge_cells(
            start_row=min_row + row_count,
            start_column=min_col,
            end_row=max_row + row_count,
            end_column=max_col,
        )
    for row in range(first_insert_row, first_insert_row + row_count):
        copy_row_style(ws, first_insert_row - 1, row, max_col=6)

# ==========================================
# 5. 写入 Excel (全范围雷达防撞机制)
# ==========================================
def process_po_to_pi():
    for folder in [INPUT_DIR, OUTPUT_DIR, "data", "templates"]:
        os.makedirs(folder, exist_ok=True)
        
    pdf_files = sorted([f for f in os.listdir(INPUT_DIR) if f.lower().endswith('.pdf') and not f.startswith('.')])
    if not pdf_files:
        print("【提示】没有发现 PDF 文件！")
        return
        
    for pdf_file in pdf_files:
        pdf_full_path = os.path.join(INPUT_DIR, pdf_file)
        po_data = extract_po_info(pdf_full_path)
        
        new_pi_no, count = generate_pi_no()
        folder_name = f"{count:02d}"
        order_folder = os.path.join(OUTPUT_DIR, folder_name)
        os.makedirs(order_folder, exist_ok=True)
        
        wb = load_workbook(TEMPLATE_PATH)
        ws = wb.active
        
        ws[CELL_PI_NO] = new_pi_no
        ws[CELL_PO_NO] = po_data["po_number"]
        ws[CELL_INVOICE_DATE] = datetime.now().strftime("%Y-%m-%d") 
        ws[CELL_BILL_TO] = po_data["bill_to"]
        ws[CELL_SHIP_TO] = po_data["deliver_to"]

        products = po_data["products"]
        # The template contains two styled detail rows (15 and 16). Reserve
        # all additional rows at once so merged payment cells move together.
        insert_product_rows(ws, START_ROW_PRODUCTS + 2, max(0, len(products) - 2))

        current_row = START_ROW_PRODUCTS
        for item_index, prod in enumerate(products, start=1):
            sku, eng_name, price = map_product_from_single_table(prod)
            qty = prod["qty"]
            amount = qty * price

            ws[f"{COL_ITEM}{current_row}"] = item_index
            ws[f"{COL_SKU}{current_row}"] = sku
            ws[f"{COL_PROD}{current_row}"] = eng_name
            ws[f"{COL_QTY}{current_row}"] = qty
            ws[f"{COL_PRICE}{current_row}"] = price
            ws[f"{COL_AMOUNT}{current_row}"] = amount
            
            current_row += 1

        total_row = current_row
        refresh_total_row(ws, START_ROW_PRODUCTS, current_row - 1, total_row)
            
        output_excel_name = f"Libernovo-Relax The Back invoice-{new_pi_no} PO{po_data['po_number']}.xlsx"
        output_excel_path = os.path.join(order_folder, output_excel_name)
        wb.save(output_excel_path)
        
        shutil.copy(pdf_full_path, os.path.join(order_folder, pdf_file))
        
        print(f"✅ 成功生成 PI，已存放至文件夹 -> [output_pi/{folder_name}]")

if __name__ == "__main__":
    process_po_to_pi()
